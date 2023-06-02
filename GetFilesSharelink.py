import requests
import re
from bs4 import BeautifulSoup
from openpyxl import Workbook


headers = {'User-agent':'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 Safari/537.36'}


with open(r"C:\Users\yuangaowei\Downloads\Share Links · NetBrain Technologies.html","r", encoding="utf-8") as f:
    content = f.read()

soup = BeautifulSoup(content, "html.parser")
links = soup.find_all("a", {"aria-label": "Visit Share Link"})

# 创建 Excel 工作簿和工作表
wb = Workbook()
ws = wb.active

# 将所有链接和它们的匹配结果写入 Excel 文件
for index, link in enumerate(links, start=1):
    # 判断链接是否符合条件
    res = requests.get(link["href"], headers=headers)
    soup1 = BeautifulSoup(res.content, "html.parser")
    finda =  soup1.find_all("a", {"aria-label": "Root Folder"})
    if finda:
        result = "Match"
    else:
        result = "No Match"
    # 将链接和匹配结果写入 Excel 文件
    ws.cell(row=index, column=1, value=link["href"])
    ws.cell(row=index, column=2, value=result)

# 保存 Excel 文件
wb.save("links.xlsx")
